#!/usr/bin/env python3
"""Build a reproducible multi-dataset OpenHermes cache-KV result workbook."""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


WORKSPACE = Path("/data/smy_data")
PRIMARY_ROOT = WORKSPACE / "local/final_results/cache-kv/openhermes_0.6+4b"
RECEIVER_ROOT = (
    WORKSPACE
    / "local/final_results/cache-kv/openhermes_lcf_projected_kv_0.6+1.5/receiver_only"
)
OUTPUT_XLSX = PRIMARY_ROOT / "openhermes_0.6+4b_multidataset_results_with_receiver_only.xlsx"
OUTPUT_CSV = PRIMARY_ROOT / "openhermes_0.6+4b_multidataset_results_with_receiver_only.csv"


@dataclass(frozen=True)
class MethodSpec:
    key: str
    label: str
    short_label: str
    model: str
    root: Path
    dataset_dirs: dict[str, str]
    raw_kv: bool = False
    baseline: bool = False
    note: str = ""


DATASETS = ("ARC-Challenge", "C-Eval", "MMLU-Redux", "OpenBookQA")
METHODS = (
    MethodSpec(
        key="lcf_fixed_quant",
        label="CacheCodec (LCF + Fixed Quant)",
        short_label="Fixed Quant",
        model="Qwen3-0.6B receiver + Qwen3-4B sharer",
        root=PRIMARY_ROOT / "lcf_jpeg",
        dataset_dirs={
            "ARC-Challenge": "ai2_arc",
            "C-Eval": "ceval",
            "MMLU-Redux": "mmlu-redux",
            "OpenBookQA": "openbookqa",
        },
        note="LCF projected-KV + fixed CacheJPEG quantization; 50 MB/s socketpair.",
    ),
    MethodSpec(
        key="lcf_adaptive_quant",
        label="CacheCodec (LCF + Adaptive Quant)",
        short_label="Adaptive Quant",
        model="Qwen3-0.6B receiver + Qwen3-4B sharer",
        root=PRIMARY_ROOT / "lcf_adaptive_quant",
        dataset_dirs={
            "ARC-Challenge": "arc_challenge",
            "C-Eval": "ceval",
            "MMLU-Redux": "mmlu_redux",
            "OpenBookQA": "openbookqa",
        },
        note="LCF projected-KV + trained adaptive quantization table; 50 MB/s socketpair.",
    ),
    MethodSpec(
        key="rosetta_raw_kv",
        label="C2C/Rosetta (Raw KV, 50 MB/s)",
        short_label="Raw KV",
        model="Qwen3-0.6B receiver + Qwen3-4B sharer",
        root=PRIMARY_ROOT / "rosetta_raw_kv_transport_50mbps",
        dataset_dirs={
            "ARC-Challenge": "arc_challenge",
            "C-Eval": "ceval",
            "MMLU-Redux": "mmlu_redux",
            "OpenBookQA": "openbookqa",
        },
        raw_kv=True,
        note="Uncompressed raw-KV transport; compression is defined as 1x and saving as 0%.",
    ),
    MethodSpec(
        key="sharer_only",
        label="Sharer-only",
        short_label="Sharer-only",
        model="Qwen3-4B-Instruct-2507",
        root=PRIMARY_ROOT / "sharer_only",
        dataset_dirs={
            "ARC-Challenge": "ai2_arc",
            "C-Eval": "ceval",
            "MMLU-Redux": "mmlu_redux",
            "OpenBookQA": "openbookqa",
        },
        baseline=True,
        note="Single-model baseline; no cache transmission.",
    ),
    MethodSpec(
        key="receiver_only",
        label="Receiver-only",
        short_label="Receiver-only",
        model="Qwen3-0.6B (reference copied from the 0.6B+1.5B result tree)",
        root=RECEIVER_ROOT,
        dataset_dirs={
            "ARC-Challenge": "arc_challenge",
            "C-Eval": "ceval",
            "MMLU-Redux": "mmlu-redux",
            "OpenBookQA": "openbookqa",
        },
        baseline=True,
        note="Receiver-only is independent of sharer size; source is the user-specified 0.6B+1.5B directory.",
    ),
)


DETAIL_FIELDS = (
    "dataset",
    "method",
    "method_key",
    "model",
    "status",
    "samples",
    "accuracy",
    "compression_ratio_x",
    "payload_to_sharer_ratio",
    "space_saving_ratio",
    "avg_sharer_cache_bytes",
    "avg_lcf_latent_kv_bytes",
    "avg_payload_bytes",
    "end_to_end_avg_ms",
    "end_to_end_p50_ms",
    "end_to_end_p95_ms",
    "avg_encode_ms",
    "avg_decode_ms",
    "avg_lcf_encode_ms",
    "avg_lcf_decode_ms",
    "avg_sender_encode_ms",
    "avg_receiver_decode_ms",
    "avg_serialize_ms",
    "avg_transmit_ms",
    "avg_deserialize_ms",
    "avg_transport_total_ms",
    "avg_bandwidth_only_transmit_ms",
    "run_timestamp",
    "summary_source",
    "performance_source",
    "note",
)


def _read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _latest(directory: Path, suffix: str) -> Path | None:
    matches = sorted(directory.glob(f"*_{suffix}.json"))
    return matches[-1] if matches else None


def _sample_count(summary: dict[str, Any], performance: dict[str, Any]) -> int | None:
    cache = summary.get("cache_transfer_statistics", {}).get("overall", {})
    if cache.get("num_samples") is not None:
        return int(cache["num_samples"])
    if performance.get("num_timed_samples") is not None:
        return int(performance["num_timed_samples"])
    subjects = summary.get("length_statistics", {}).get("subjects", {})
    counts = [value.get("total_samples") for value in subjects.values()]
    return int(sum(value for value in counts if value is not None)) if counts else None


def collect_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for dataset in DATASETS:
        for method in METHODS:
            directory = method.root / method.dataset_dirs[dataset]
            summary_path = _latest(directory, "summary")
            performance_path = _latest(directory, "performance")
            if summary_path is None:
                rows.append(
                    {
                        "dataset": dataset,
                        "method": method.label,
                        "method_key": method.key,
                        "model": method.model,
                        "status": "Missing result",
                        "samples": None,
                        "accuracy": None,
                        "compression_ratio_x": None,
                        "payload_to_sharer_ratio": None,
                        "space_saving_ratio": None,
                        "run_timestamp": None,
                        "summary_source": str(directory),
                        "performance_source": "",
                        "note": method.note,
                    }
                )
                continue

            summary = _read_json(summary_path)
            performance = _read_json(performance_path) if performance_path else summary.get("performance", {})
            cache = summary.get("cache_transfer_statistics", {}).get("overall", {})
            timestamp_match = re.search(r"_(\d{14})_summary\.json$", summary_path.name)
            row: dict[str, Any] = {
                "dataset": dataset,
                "method": method.label,
                "method_key": method.key,
                "model": method.model,
                "status": "Available",
                "samples": _sample_count(summary, performance),
                "accuracy": summary.get("overall_accuracy"),
                "compression_ratio_x": cache.get("aggregate_sharer_to_payload_compression_ratio"),
                "payload_to_sharer_ratio": cache.get("aggregate_payload_to_sharer_ratio"),
                "space_saving_ratio": cache.get("aggregate_space_saving_ratio"),
                "avg_sharer_cache_bytes": cache.get("avg_sharer_cache_bytes"),
                "avg_lcf_latent_kv_bytes": cache.get("avg_lcf_latent_kv_bytes"),
                "avg_payload_bytes": cache.get("avg_payload_bytes"),
                "run_timestamp": timestamp_match.group(1) if timestamp_match else None,
                "summary_source": str(summary_path),
                "performance_source": str(performance_path) if performance_path else "embedded in summary",
                "note": method.note,
            }
            for key in DETAIL_FIELDS:
                if key not in row and key in performance:
                    row[key] = performance[key]
            if method.raw_kv:
                row["compression_ratio_x"] = 1.0
                row["payload_to_sharer_ratio"] = 1.0
                row["space_saving_ratio"] = 0.0
            rows.append(row)
    return rows


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
MISSING_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_GRAY = Side(style="thin", color="B7B7B7")


def _style_header(cell, *, primary: bool = True) -> None:
    cell.fill = HEADER_FILL if primary else SUBHEADER_FILL
    cell.font = Font(color="FFFFFF" if primary else "000000", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=THIN_GRAY)


def _autosize(ws, maximum: int = 42) -> None:
    for index, cells in enumerate(ws.columns, start=1):
        width = 0
        for cell in cells:
            if cell.value is not None:
                parts = str(cell.value).splitlines() or [""]
                width = max(width, max(len(part) for part in parts))
        ws.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), maximum)


def _write_main_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = workbook.active
    ws.title = "论文主表"
    headers = (
        "数据集",
        "方法",
        "正确率",
        "压缩倍数",
        "Payload / Raw KV",
        "端到端平均时间",
        "传输平均时间",
    )
    ws.append(headers)
    for cell in ws[1]:
        _style_header(cell)
    method_order = {
        "lcf_fixed_quant": 0,
        "lcf_adaptive_quant": 1,
        "receiver_only": 2,
        "sharer_only": 3,
        "rosetta_raw_kv": 4,
    }
    main_labels = {
        "lcf_fixed_quant": "LCF-Projected-KV + CacheJPEG",
        "lcf_adaptive_quant": "LCF-Projected-KV + Adaptive Quant",
        "receiver_only": "Receiver-only",
        "sharer_only": "Sharer-only",
        "rosetta_raw_kv": "C2C/Rosetta raw-KV（50 MB/s）",
    }
    dataset_order = {dataset: index for index, dataset in enumerate(DATASETS)}
    ordered_rows = sorted(
        rows,
        key=lambda row: (dataset_order[row["dataset"]], method_order[row["method_key"]]),
    )
    previous_dataset = None
    for row in ordered_rows:
        available = row["status"] == "Available"
        dataset_cell = row["dataset"] if row["dataset"] != previous_dataset else ""
        previous_dataset = row["dataset"]
        ws.append(
            (
                dataset_cell,
                main_labels[row["method_key"]],
                row.get("accuracy") if available else "未测试",
                row.get("compression_ratio_x") if available else None,
                row.get("payload_to_sharer_ratio") if available else None,
                row.get("end_to_end_avg_ms") if available else None,
                row.get("avg_transmit_ms") if available else None,
            )
        )
        if not available:
            for cell in ws[ws.max_row]:
                cell.fill = MISSING_FILL
    for cell in ws["C"][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"
    for cell in ws["D"][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00"×"'
        elif cell.value is None:
            cell.value = "—"
    for cell in ws["E"][1:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"
        elif cell.value is None:
            cell.value = "—"
    for column in ("F", "G"):
        for cell in ws[column][1:]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00 "ms"'
            elif cell.value is None:
                cell.value = "—"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _write_wide_sheet(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    title: str,
    metrics: tuple[tuple[str, str, str], ...],
) -> None:
    ws = workbook.create_sheet(title)
    ws.cell(1, 1, "数据集")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    _style_header(ws.cell(1, 1))
    lookup = {(row["dataset"], row["method_key"]): row for row in rows}
    column = 2
    for method in METHODS:
        start = column
        for _key, label, _fmt in metrics:
            ws.cell(2, column, label)
            _style_header(ws.cell(2, column), primary=False)
            column += 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=column - 1)
        ws.cell(1, start, method.short_label)
        _style_header(ws.cell(1, start))
    for dataset_index, dataset in enumerate(DATASETS, start=3):
        ws.cell(dataset_index, 1, dataset)
        column = 2
        for method in METHODS:
            row = lookup[(dataset, method.key)]
            for key, _label, number_format in metrics:
                cell = ws.cell(dataset_index, column, row.get(key))
                cell.number_format = number_format
                if row["status"] != "Available":
                    cell.fill = MISSING_FILL
                column += 1
    ws.freeze_panes = "B3"
    _autosize(ws, maximum=20)


def _write_detail_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("完整指标")
    ws.append(DETAIL_FIELDS)
    for cell in ws[1]:
        _style_header(cell)
    for row in rows:
        ws.append([row.get(field) for field in DETAIL_FIELDS])
    percent_fields = {"accuracy", "payload_to_sharer_ratio", "space_saving_ratio"}
    for column_index, field in enumerate(DETAIL_FIELDS, start=1):
        if field in percent_fields:
            for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                cell[0].number_format = "0.0000%"
        elif field.endswith("_ms") or field == "compression_ratio_x":
            for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                cell[0].number_format = "0.000000"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _write_readme_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = workbook.create_sheet("口径与来源")
    content = (
        ("项目", "说明"),
        ("主结果目录", str(PRIMARY_ROOT)),
        ("Receiver-only 来源", str(RECEIVER_ROOT)),
        ("数据集", ", ".join(DATASETS)),
        ("结果选择", "每个方法/数据集目录选择文件名时间戳最新的一组 summary/performance JSON。"),
        ("正确率", "summary.overall_accuracy；Excel 显示为百分比，CSV 保留 [0,1] 原值。"),
        ("压缩倍数", "total_sharer_cache_bytes / total_payload_bytes，采用 summary 中 aggregate_sharer_to_payload_compression_ratio。"),
        ("空间节省率", "1 - total_payload_bytes / total_sharer_cache_bytes。"),
        ("传输时间", "performance.avg_transmit_ms，仅传输调用耗时。"),
        ("传输总计", "performance.avg_transport_total_ms，包含序列化、传输和反序列化。"),
        ("端到端时间", "performance.end_to_end_avg_ms；同时保留 P50/P95 和总秒数。"),
        ("Raw-KV 压缩率", "原始 KV 无压缩，按定义记录 1x、payload/raw=100%、空间节省=0%；源 JSON 未记录平均 payload 字节，故留空。"),
        ("单模型基线", "Sharer-only/Receiver-only 不发生 KV 传输，压缩和传输字段留空，不记为 0。"),
        ("缺失结果", "LCF + Adaptive Quant 的 C-Eval 目录当前没有 summary JSON，在表中显式标为 Missing result。"),
        ("可用组合数", f"{sum(row['status'] == 'Available' for row in rows)} / {len(rows)}"),
    )
    for row in content:
        ws.append(row)
    for cell in ws[1]:
        _style_header(cell)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_csv(rows: list[dict[str, Any]]) -> None:
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def build_workbook(rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    _write_main_sheet(workbook, rows)
    _write_wide_sheet(
        workbook,
        rows,
        "按数据集-质量压缩",
        (
            ("accuracy", "正确率", "0.00%"),
            ("compression_ratio_x", "压缩倍数", "0.000"),
            ("space_saving_ratio", "空间节省率", "0.00%"),
        ),
    )
    _write_wide_sheet(
        workbook,
        rows,
        "按数据集-时延",
        (
            ("end_to_end_avg_ms", "端到端 (ms)", "0.000"),
            ("avg_transmit_ms", "传输 (ms)", "0.000"),
            ("avg_transport_total_ms", "传输总计 (ms)", "0.000"),
        ),
    )
    _write_detail_sheet(workbook, rows)
    _write_readme_sheet(workbook, rows)
    workbook.save(OUTPUT_XLSX)


def validate_outputs(expected_rows: int) -> None:
    workbook = load_workbook(OUTPUT_XLSX, read_only=True, data_only=True)
    assert workbook.sheetnames == [
        "论文主表",
        "按数据集-质量压缩",
        "按数据集-时延",
        "完整指标",
        "口径与来源",
    ]
    assert workbook["论文主表"].max_row == expected_rows + 1
    assert workbook["完整指标"].max_row == expected_rows + 1
    workbook.close()


def main() -> None:
    rows = collect_rows()
    write_csv(rows)
    build_workbook(rows)
    validate_outputs(len(rows))
    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_CSV}")
    print(f"Available result combinations: {sum(row['status'] == 'Available' for row in rows)}/{len(rows)}")


if __name__ == "__main__":
    main()
