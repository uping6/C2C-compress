"""Build an accuracy/latency/transport comparison from saved evaluation artifacts."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


ROOT = Path("/data/smy_data/local/final_results/openhermes_lcf_projected_kv")
OUT_DIR = ROOT

RUNS = [
    (
        "MMLU-Redux",
        "LCF-Projected-KV + CacheJPEG (parallel)",
        ROOT / "mmlu/parallel",
        "compressed",
    ),
    ("MMLU-Redux", "Receiver-only (Qwen3-0.6B)", ROOT / "receiver_only/mmlu-redux", "single"),
    ("MMLU-Redux", "Sharer-only (Qwen2.5-1.5B-Instruct)", ROOT / "sharer_only/mmlu-redux", "single"),
    ("MMLU-Redux", "C2C/Rosetta raw-KV transport (50 MB/s)", ROOT / "rosetta_raw_kv_transport_50mbps/mmlu-redux", "raw"),
    (
        "OpenBookQA",
        "LCF-Projected-KV + CacheJPEG (parallel)",
        ROOT / "openbookqa/parallel",
        "compressed",
    ),
    ("OpenBookQA", "Receiver-only (Qwen3-0.6B)", ROOT / "receiver_only/openbookqa", "single"),
    ("OpenBookQA", "Sharer-only (Qwen2.5-1.5B-Instruct)", ROOT / "sharer_only/openbookqa", "single"),
    ("OpenBookQA", "C2C/Rosetta raw-KV transport (50 MB/s)", ROOT / "rosetta_raw_kv_transport_50mbps/openbookqa", "raw"),
]


def newest(directory: Path, suffix: str) -> Path:
    files = list(directory.glob(f"*{suffix}.json"))
    if not files:
        raise FileNotFoundError(f"No *{suffix}.json under {directory}")
    return max(files, key=lambda item: item.stat().st_mtime)


def mean(rows: list[dict[str, Any]], field: str) -> float | None:
    values = [float(row[field]) for row in rows if row.get(field) is not None]
    return sum(values) / len(values) if values else None


def build_row(dataset: str, method: str, directory: Path, kind: str) -> dict[str, Any]:
    summary_path = newest(directory, "summary")
    performance_path = newest(directory, "performance")
    length_path = newest(directory, "length")
    summary = json.loads(summary_path.read_text())
    performance = json.loads(performance_path.read_text())
    lengths = json.loads(length_path.read_text())

    compression_factor: float | None = None
    payload_ratio: float | None = None
    if kind == "compressed":
        compression_factor = mean(lengths, "sharer_to_payload_compression_ratio")
        payload_ratio = mean(lengths, "payload_to_sharer_ratio")
    elif kind == "raw":
        # The transmitted object is the uncompressed sharer KV; pickle framing
        # is transport overhead, not a compression method.
        compression_factor = 1.0
        payload_ratio = 1.0

    return {
        "数据集": dataset,
        "方法": method,
        "样本数": len(lengths),
        "正确率": float(summary["overall_accuracy"]),
        "压缩倍数(raw_KV/payload)": compression_factor,
        "payload占raw_KV比例": payload_ratio,
        "平均端到端时间(ms)": performance.get("end_to_end_avg_ms"),
        "平均传输时间(ms)": performance.get("avg_transmit_ms"),
        "平均传输总时间(ms)": performance.get("avg_transport_total_ms"),
        "平均payload(bytes)": mean(lengths, "transport_payload_bytes") or mean(lengths, "payload_bytes"),
        "summary来源": str(summary_path),
        "performance来源": str(performance_path),
        "说明": (
            "原始KV无压缩；50 MB/s 串行 socketpair transport"
            if kind == "raw"
            else "单模型基线：无 sharer KV 传输"
            if kind == "single"
            else "LCF projected-KV + CacheJPEG；parallel 流水线"
        ),
    }


def main() -> None:
    rows = [build_row(*run) for run in RUNS]
    csv_path = OUT_DIR / "mmlu_openbookqa_method_comparison.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    xlsx_path = OUT_DIR / "mmlu_openbookqa_method_comparison.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        # CSV is the authoritative machine-readable output.  Environments that
        # do not ship an Excel writer can convert it with LibreOffice.
        print(csv_path)
        print("openpyxl is unavailable; convert the CSV with LibreOffice to create XLSX.")
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "方法对比"
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[key] for key in headers])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in ("正确率", "payload占raw_KV比例"):
        idx = headers.index(column) + 1
        for cell in sheet.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for item in cell:
                if item.value is not None:
                    item.number_format = "0.00%"
    for column in (
        "压缩倍数(raw_KV/payload)",
        "平均端到端时间(ms)",
        "平均传输时间(ms)",
        "平均传输总时间(ms)",
        "平均payload(bytes)",
    ):
        idx = headers.index(column) + 1
        for cell in sheet.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for item in cell:
                if item.value is not None:
                    item.number_format = "0.00"
    for idx, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(len(header) + 2, 16), 38)
    sheet.freeze_panes = "A2"

    notes = workbook.create_sheet("口径说明")
    notes.append(["字段", "说明"])
    notes.append(["压缩倍数", "raw sharer KV bytes / 实际 payload bytes；越大表示压缩越强。"])
    notes.append(["payload占raw_KV比例", "实际 payload / raw sharer KV；越小表示压缩越强。"])
    notes.append(["raw-KV transport", "无压缩，按 1.00× / 100% 标记；pickle/socket framing 不计为压缩。"])
    notes.append(["单模型基线", "没有 sharer KV 或传输，压缩率与传输时间不适用。"])
    notes.append(["时间", "均为每样本平均端到端墙钟时间或传输阶段时间。"])
    for cell in notes[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 100
    for row in notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(xlsx_path)
    print(csv_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
